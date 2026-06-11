from __future__ import annotations

from pathlib import Path
import shutil

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


PROJECT = Path(__file__).resolve().parents[1]
DATE = "2026-06-07"
SRC_DATE = "2026-06-05"

RESULTS = PROJECT / "results"
OUT = PROJECT / "\u6295\u7a3f" / "manuscript_tables"
MAIN = OUT / "main_tables"
SUPP = OUT / "supplementary_tables"
CSV_OUT = OUT / "csv_sources"


def read_csv(path: Path, **kwargs) -> pd.DataFrame:
    return pd.read_csv(path, low_memory=False, **kwargs)


def pct_col(frame: pd.DataFrame, col: str, out: str | None = None) -> None:
    if col in frame.columns:
        frame[out or f"{col}_pct"] = frame[col].map(lambda x: "" if pd.isna(x) else round(float(x) * 100, 1))


def tidy_percent(value) -> str:
    if pd.isna(value) or value == "":
        return ""
    return f"{float(value):.1f}%"


def tidy_float(value, ndigits: int = 2) -> str:
    if pd.isna(value) or value == "":
        return ""
    return f"{float(value):.{ndigits}f}"


def safe_sheet(name: str) -> str:
    return name[:31]


def style_workbook(path: Path) -> None:
    from openpyxl import load_workbook

    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="EAF1F7")
    header_font = Font(bold=True, color="243447")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        for col_idx, column_cells in enumerate(ws.columns, start=1):
            max_len = 0
            for cell in column_cells[:200]:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, min(len(value), 60))
            ws.column_dimensions[get_column_letter(col_idx)].width = max(10, min(max_len + 2, 38))
    wb.save(path)


def write_workbook(path: Path, sheets: dict[str, pd.DataFrame]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        used: set[str] = set()
        for name, frame in sheets.items():
            sheet = safe_sheet(name)
            base = sheet
            i = 2
            while sheet in used:
                suffix = f"_{i}"
                sheet = safe_sheet(base[: 31 - len(suffix)] + suffix)
                i += 1
            used.add(sheet)
            frame.to_excel(writer, index=False, sheet_name=sheet)
    style_workbook(path)


def write_csv_copy(frame: pd.DataFrame, name: str) -> Path:
    CSV_OUT.mkdir(parents=True, exist_ok=True)
    path = CSV_OUT / name
    frame.to_csv(path, index=False, encoding="utf-8-sig")
    return path


def table1() -> tuple[Path, dict[str, pd.DataFrame]]:
    cohort = read_csv(RESULTS / "07_figure_data" / f"fig1a_dataset_readiness_{SRC_DATE}.csv")
    perf = read_csv(RESULTS / "02_unified_ipm_mem_prediction" / f"ipm_mem_unified_summary_by_drug_{SRC_DATE}.csv")

    cohort_display = cohort[
        [
            "display_label",
            "cohort_role",
            "isolates",
            "assemblies_present",
            "ipm_mic_and_sir",
            "mem_mic_and_sir",
        ]
    ].rename(
        columns={
            "display_label": "Cohort",
            "cohort_role": "Role",
            "isolates": "Genomes/isolates",
            "assemblies_present": "Assemblies/features",
            "ipm_mic_and_sir": "IPM MIC/SIR",
            "mem_mic_and_sir": "MEM MIC/SIR",
        }
    )

    perf_display = perf.copy()
    perf_display = perf_display[
        [
            "cohort",
            "drug",
            "n",
            "n_s",
            "n_ns",
            "mic_ea_pm1",
            "mic_mae",
            "mic_bias",
            "sns_ca",
            "sns_vme",
            "sns_me",
            "gate_auc",
        ]
    ].rename(
        columns={
            "cohort": "Locked evaluation cohort",
            "drug": "Drug",
            "n": "n",
            "n_s": "Susceptible n",
            "n_ns": "Non-susceptible n",
            "mic_ea_pm1": "EA +/-1 dilution",
            "mic_mae": "MAE, log2 dilution",
            "mic_bias": "Bias, log2 dilution",
            "sns_ca": "Categorical agreement",
            "sns_vme": "VME rate (false S / true NS)",
            "sns_me": "ME rate (false NS / true S)",
            "gate_auc": "Gate AUC",
        }
    )
    for col in ["EA +/-1 dilution", "Categorical agreement", "VME rate (false S / true NS)", "ME rate (false NS / true S)"]:
        perf_display[col] = perf_display[col].map(lambda x: tidy_percent(float(x) * 100))
    for col in ["MAE, log2 dilution", "Bias, log2 dilution", "Gate AUC"]:
        perf_display[col] = perf_display[col].map(lambda x: tidy_float(x, 2))

    path = MAIN / f"Table_1_cohort_composition_and_locked_model_evaluation_{DATE}.xlsx"
    sheets = {
        "cohort_composition_display": cohort_display,
        "locked_evaluation_display": perf_display,
        "cohort_composition_source": cohort,
        "locked_evaluation_source": perf,
    }
    write_workbook(path, sheets)
    write_csv_copy(cohort_display, f"Table_1A_cohort_composition_display_{DATE}.csv")
    write_csv_copy(perf_display, f"Table_1B_locked_evaluation_display_{DATE}.csv")
    return path, sheets


def table2() -> tuple[Path, dict[str, pd.DataFrame]]:
    raw = read_csv(RESULTS / "10_clinical_warning_rules" / f"clinical_warning_rule_table_{SRC_DATE}.csv")
    display = raw[
        [
            "drug",
            "mechanism_subtype_high_confidence",
            "n",
            "n_s",
            "n_ns",
            "within_1_dilution",
            "mae_log2",
            "bias_log2",
            "categorical_agreement",
            "very_major_error_rate",
            "major_error_rate",
            "large_error_gt2",
            "clinical_rule",
            "rule_reason",
        ]
    ].rename(
        columns={
            "drug": "Drug",
            "mechanism_subtype_high_confidence": "High-confidence mechanism subtype",
            "n": "n",
            "n_s": "Susceptible n",
            "n_ns": "Non-susceptible n",
            "within_1_dilution": "EA +/-1 dilution",
            "mae_log2": "MAE, log2 dilution",
            "bias_log2": "Bias, log2 dilution",
            "categorical_agreement": "Categorical agreement",
            "very_major_error_rate": "VME rate (false S / true NS)",
            "major_error_rate": "ME rate (false NS / true S)",
            "large_error_gt2": "Large error >2 log2",
            "clinical_rule": "Clinical interpretation rule",
            "rule_reason": "Rule rationale",
        }
    )
    for col in ["EA +/-1 dilution", "Categorical agreement", "VME rate (false S / true NS)", "ME rate (false NS / true S)", "Large error >2 log2"]:
        display[col] = display[col].map(lambda x: tidy_percent(float(x) * 100 if abs(float(x)) <= 1 else x))
    for col in ["MAE, log2 dilution", "Bias, log2 dilution"]:
        display[col] = display[col].map(lambda x: tidy_float(x, 2))

    path = MAIN / f"Table_2_high_confidence_clinical_interpretation_rules_{DATE}.xlsx"
    sheets = {
        "clinical_rules_display": display,
        "clinical_rules_source": raw,
    }
    write_workbook(path, sheets)
    write_csv_copy(display, f"Table_2_clinical_interpretation_rules_display_{DATE}.csv")
    return path, sheets


def supplementary_tables() -> list[dict[str, object]]:
    paths: list[dict[str, object]] = []

    def add(table_id: str, title: str, sheets: dict[str, pd.DataFrame]) -> None:
        path = SUPP / f"{table_id}_{title.replace(' ', '_').replace('/', '_')}_{DATE}.xlsx"
        write_workbook(path, sheets)
        paths.append({"table_id": table_id, "title": title, "path": path, "sheets": list(sheets)})

    add(
        "Supplementary_Table_S01",
        "model_policy_and_endpoint_definitions",
        {
            "model_policy": read_csv(PROJECT / "models" / "ipm_mem_unified_public_only" / f"ipm_mem_unified_public_only_policy_{SRC_DATE}.csv"),
            "training_apparent_summary": read_csv(PROJECT / "models" / "ipm_mem_unified_public_only" / f"ipm_mem_unified_public_only_training_apparent_summary_{SRC_DATE}.csv"),
        },
    )
    add(
        "Supplementary_Table_S02",
        "full_locked_evaluation_by_cohort_and_drug",
        {
            "summary_by_drug": read_csv(RESULTS / "02_unified_ipm_mem_prediction" / f"ipm_mem_unified_summary_by_drug_{SRC_DATE}.csv"),
            "summary_overall": read_csv(RESULTS / "02_unified_ipm_mem_prediction" / f"ipm_mem_unified_summary_overall_{SRC_DATE}.csv"),
            "prediction_errors": read_csv(RESULTS / "02_unified_ipm_mem_prediction" / f"ipm_mem_unified_prediction_errors_{SRC_DATE}.tsv", sep="\t"),
        },
    )
    add(
        "Supplementary_Table_S03",
        "local_paired_ipm_mem_phenotype_group_metrics",
        {
            "pair_group_metrics": read_csv(RESULTS / "02_unified_ipm_mem_prediction" / f"ipm_mem_unified_local_pair_group_summary_{SRC_DATE}.csv"),
            "pair_group_counts": read_csv(RESULTS / "07_figure_data" / f"fig1c_local_ipm_mem_pair_groups_{SRC_DATE}.csv"),
        },
    )
    add(
        "Supplementary_Table_S04",
        "first_pass_strict_mechanism_evidence",
        {
            "strict_evidence": read_csv(RESULTS / "03_mechanism_predictability" / f"mechanism_evidence_strict_first_pass_{SRC_DATE}.tsv", sep="\t"),
            "strict_predictability": read_csv(RESULTS / "03_mechanism_predictability" / f"subtype_predictability_summary_by_drug_{SRC_DATE}.csv"),
        },
    )
    add(
        "Supplementary_Table_S05",
        "high_confidence_mechanism_evidence_per_isolate",
        {
            "high_confidence_evidence": read_csv(RESULTS / "08_deep_mechanism_annotation" / f"local_deep_mechanism_evidence_{SRC_DATE}.tsv", sep="\t"),
            "prediction_with_mechanism": read_csv(RESULTS / "08_deep_mechanism_annotation" / f"local_prediction_errors_with_deep_mechanism_{SRC_DATE}.tsv", sep="\t"),
        },
    )
    add(
        "Supplementary_Table_S06",
        "target_gene_variant_summary",
        {
            "target_gene_variant_summary": read_csv(RESULTS / "08_deep_mechanism_annotation" / f"target_gene_variant_summary_{SRC_DATE}.tsv", sep="\t"),
            "target_variant_rows": read_csv(RESULTS / "08_deep_mechanism_annotation" / f"target_variant_rows_{SRC_DATE}.tsv", sep="\t"),
        },
    )
    add(
        "Supplementary_Table_S07",
        "high_confidence_paired_ipm_mem_statistical_tests",
        {
            "paired_tests": read_csv(RESULTS / "09_high_confidence_statistics" / f"high_confidence_paired_ipm_mem_tests_{SRC_DATE}.csv"),
            "global_tests": read_csv(RESULTS / "09_high_confidence_statistics" / f"high_confidence_global_tests_{SRC_DATE}.csv"),
            "predictability_summary": read_csv(RESULTS / "09_high_confidence_statistics" / f"high_confidence_predictability_summary_{SRC_DATE}.csv"),
        },
    )
    add(
        "Supplementary_Table_S08",
        "mlst_and_st_sensitivity",
        {
            "mlst_calls": read_csv(RESULTS / "05_mlst" / f"local_mlst_pubmlst_exact_{SRC_DATE}.tsv", sep="\t"),
            "st_counts": read_csv(RESULTS / "06_st_sensitivity" / f"local_st_counts_{SRC_DATE}.csv"),
            "leave_one_st_out": read_csv(RESULTS / "06_st_sensitivity" / f"local_leave_one_st_out_sensitivity_{SRC_DATE}.csv"),
            "st_adjusted_models": read_csv(RESULTS / "06_st_sensitivity" / f"local_st_adjusted_models_{SRC_DATE}.csv"),
            "paired_error_by_st": read_csv(RESULTS / "06_st_sensitivity" / f"local_paired_error_difference_by_st_{SRC_DATE}.csv"),
        },
    )
    add(
        "Supplementary_Table_S09",
        "clinical_warning_rule_full_table",
        {
            "clinical_warning_rules": read_csv(RESULTS / "10_clinical_warning_rules" / f"clinical_warning_rule_table_{SRC_DATE}.csv"),
            "rule_level_performance": read_csv(RESULTS / "10_clinical_warning_rules" / f"clinical_rule_level_performance_{SRC_DATE}.csv"),
            "mem_warning_enrichment": read_csv(RESULTS / "10_clinical_warning_rules" / f"mem_warning_enrichment_tests_{SRC_DATE}.csv"),
        },
    )
    add(
        "Supplementary_Table_S10",
        "false_susceptible_cases",
        {
            "false_susceptible_cases": read_csv(RESULTS / "10_clinical_warning_rules" / f"false_susceptible_cases_by_warning_rule_{SRC_DATE}.csv"),
            "top_error_cases": read_csv(RESULTS / "09_high_confidence_statistics" / f"high_confidence_top_error_cases_{SRC_DATE}.csv"),
        },
    )
    add(
        "Supplementary_Table_S11",
        "calibration_and_breakpoint_distance_summaries",
        {
            "calibration_summary": read_csv(RESULTS / "11_enriched_clinical_error_analyses" / f"calibration_summary_by_high_confidence_subtype_{SRC_DATE}.csv"),
            "calibration_bins": read_csv(RESULTS / "11_enriched_clinical_error_analyses" / f"calibration_by_probability_bin_{SRC_DATE}.csv"),
            "susceptible_call_risk": read_csv(RESULTS / "11_enriched_clinical_error_analyses" / f"susceptible_call_risk_by_confidence_{SRC_DATE}.csv"),
            "breakpoint_zone_errors": read_csv(RESULTS / "11_enriched_clinical_error_analyses" / f"breakpoint_zone_error_summary_{SRC_DATE}.csv"),
            "phenotype_combo_errors": read_csv(RESULTS / "11_enriched_clinical_error_analyses" / f"phenotype_combo_paired_error_summary_{SRC_DATE}.csv"),
        },
    )
    add(
        "Supplementary_Table_S12",
        "error_taxonomy_and_safety_gate_evaluation",
        {
            "error_taxonomy_summary": read_csv(RESULTS / "11_enriched_clinical_error_analyses" / f"error_taxonomy_summary_{SRC_DATE}.csv"),
            "error_taxonomy_rows": read_csv(RESULTS / "11_enriched_clinical_error_analyses" / f"error_taxonomy_rows_{SRC_DATE}.tsv", sep="\t"),
            "safety_gate_summary": read_csv(RESULTS / "11_enriched_clinical_error_analyses" / f"safety_gate_evaluation_summary_{SRC_DATE}.csv"),
        },
    )
    return paths


def build_all_workbook(main_tables: list[tuple[str, Path, dict[str, pd.DataFrame]]], supp_rows: list[dict[str, object]]) -> Path:
    path = OUT / f"IPM-GPT_all_manuscript_tables_{DATE}.xlsx"
    sheets: dict[str, pd.DataFrame] = {}
    for table_id, _path, frames in main_tables:
        for sheet_name, frame in frames.items():
            sheets[f"{table_id}_{sheet_name}"] = frame
    for row in supp_rows:
        table_id = str(row["table_id"]).replace("Supplementary_Table_", "ST")
        for sheet in row["sheets"]:
            # Load from already written file to avoid carrying very large raw frames in memory longer.
            pass
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for table_id, _path, frames in main_tables:
            for sheet_name, frame in frames.items():
                frame.to_excel(writer, index=False, sheet_name=safe_sheet(f"{table_id}_{sheet_name}"))
        for row in supp_rows:
            xls = pd.ExcelFile(row["path"])
            for sheet_name in xls.sheet_names:
                frame = pd.read_excel(xls, sheet_name=sheet_name)
                frame.to_excel(writer, index=False, sheet_name=safe_sheet(f"{row['table_id'].replace('Supplementary_Table_', 'ST')}_{sheet_name}"))
    style_workbook(path)
    return path


def write_manifest(main_paths: list[dict[str, object]], supp_paths: list[dict[str, object]], all_workbook: Path) -> Path:
    rows = []
    rows.extend(main_paths)
    for row in supp_paths:
        rows.append(
            {
                "table_id": row["table_id"],
                "title": row["title"],
                "destination": "supplementary",
                "file": str(Path(row["path"]).relative_to(OUT)).replace("\\", "/"),
                "sheets": "; ".join(row["sheets"]),
            }
        )
    rows.append(
        {
            "table_id": "all_tables_workbook",
            "title": "All manuscript tables combined workbook",
            "destination": "main_and_supplementary",
            "file": all_workbook.name,
            "sheets": "all",
        }
    )
    manifest = pd.DataFrame(rows)
    path = OUT / f"manuscript_table_manifest_{DATE}.csv"
    manifest.to_csv(path, index=False, encoding="utf-8-sig")
    return path


def clean_output() -> None:
    if OUT.exists():
        shutil.rmtree(OUT)
    MAIN.mkdir(parents=True, exist_ok=True)
    SUPP.mkdir(parents=True, exist_ok=True)
    CSV_OUT.mkdir(parents=True, exist_ok=True)


def main() -> None:
    clean_output()
    t1_path, t1_sheets = table1()
    t2_path, t2_sheets = table2()
    main_manifest = [
        {
            "table_id": "Table_1",
            "title": "Cohort composition and locked model evaluation",
            "destination": "main",
            "file": str(t1_path.relative_to(OUT)).replace("\\", "/"),
            "sheets": "; ".join(t1_sheets),
        },
        {
            "table_id": "Table_2",
            "title": "High-confidence mechanism strata and clinical interpretation",
            "destination": "main",
            "file": str(t2_path.relative_to(OUT)).replace("\\", "/"),
            "sheets": "; ".join(t2_sheets),
        },
    ]
    supp = supplementary_tables()
    all_workbook = build_all_workbook(
        [("T1", t1_path, t1_sheets), ("T2", t2_path, t2_sheets)],
        supp,
    )
    manifest = write_manifest(main_manifest, supp, all_workbook)
    print(f"Wrote manuscript tables to {OUT}")
    print(f"Main tables: {MAIN}")
    print(f"Supplementary tables: {SUPP}")
    print(f"Manifest: {manifest}")
    print(f"All-tables workbook: {all_workbook}")


if __name__ == "__main__":
    main()
